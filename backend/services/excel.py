"""Excel report writer.

Sheet 1 "Theme Analysis" holds exactly the six requested columns and nothing
else, so the file can be handed straight to the business. Sheets 2 and 3 carry
the supporting detail and a theme/issue rollup.

`build_result_rows` is shared with the API, so the table rendered in the
browser and the workbook can never disagree.
"""

from __future__ import annotations

import re
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from backend.core.logging_config import get_logger
from backend.schemas.common import CallStatus
from backend.schemas.job import CallRecord, JobRecord, ResultRow

logger = get_logger(__name__)

REPORT_COLUMNS = [
    "Sr. No",
    "FileName",
    "Theme",
    "Specific Issue for the call",
    "Transcription",
    "AI Reasoning",
]

# ResultRow attribute for each column above, in the same order. The API sends
# both lists so the frontend table is driven by the report definition rather
# than by hardcoded headings.
REPORT_FIELDS = [
    "sr_no",
    "file_name",
    "theme",
    "specific_issue",
    "transcription",
    "ai_reasoning",
]

# Excel refuses anything longer in a single cell.
MAX_CELL_CHARS = 32_000
# Control characters openpyxl cannot serialise.
_ILLEGAL = re.compile(r"[\000-\010\013\014\016-\037]")

_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_FAILED_FILL = PatternFill("solid", fgColor="FCE4E4")
_THIN = Side(style="thin", color="D0D0D0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
_CENTER = Alignment(horizontal="center", vertical="top")
_RTL = Alignment(horizontal="right", vertical="top", wrap_text=True, readingOrder=2)


def clean_cell(value: str | None) -> str:
    """Strip control characters and clip to Excel's per-cell limit."""
    if not value:
        return ""
    text = _ILLEGAL.sub("", str(value))
    if len(text) > MAX_CELL_CHARS:
        text = text[: MAX_CELL_CHARS - 40] + "\n... [truncated for Excel]"
    return text


# --------------------------------------------------------------------------
# Job -> rows
# --------------------------------------------------------------------------
def build_result_rows(job: JobRecord) -> list[ResultRow]:
    return [_row_for(call) for call in sorted(job.calls, key=lambda c: c.sr_no)]


def _row_for(call: CallRecord) -> ResultRow:
    classification = call.classification

    if call.status is CallStatus.FAILED:
        theme = "PROCESSING FAILED"
        issue = f"Failed at stage: {call.failed_stage or 'unknown'}"
        reasoning = call.error or "Unknown error"
        reason = ""
    elif classification is None:
        theme = "PENDING"
        issue = "Not yet classified"
        reasoning = ""
        reason = ""
    else:
        theme = classification.theme
        issue = classification.issue
        reasoning = classification.to_reasoning_cell()
        reason = classification.reason

    return ResultRow(
        sr_no=call.sr_no,
        file_name=call.filename,
        theme=theme,
        specific_issue=issue,
        transcription=call.final_transcript,
        ai_reasoning=reasoning,
        call_id=call.id,
        status=call.status,
        reason_for_issue=reason,
        confidence=classification.confidence if classification else 0.0,
        language=call.transcription.language if call.transcription else "unknown",
        duration_seconds=(
            call.transcription.duration_seconds if call.transcription else 0.0
        ),
        original_transcription=call.original_transcript,
        evidence=classification.evidence if classification else [],
        error=call.error,
    )


# --------------------------------------------------------------------------
# Workbook
# --------------------------------------------------------------------------
def write_report(job: JobRecord, output_path: Path) -> Path:
    """Blocking - call it from a thread (job_manager uses asyncio.to_thread)."""
    rows = build_result_rows(job)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    _write_main_sheet(workbook.active, rows)
    _write_details_sheet(workbook.create_sheet("Details"), rows)
    _write_summary_sheet(workbook.create_sheet("Summary"), job, rows)

    workbook.save(output_path)
    logger.info("Report written: %s (%d rows)", output_path, len(rows))
    return output_path


def _write_main_sheet(sheet: Worksheet, rows: list[ResultRow]) -> None:
    sheet.title = "Theme Analysis"
    sheet.append(REPORT_COLUMNS)

    for row in rows:
        sheet.append(
            [
                row.sr_no,
                clean_cell(row.file_name),
                clean_cell(row.theme),
                clean_cell(row.specific_issue),
                clean_cell(row.transcription),
                clean_cell(row.ai_reasoning),
            ]
        )
        if row.status is CallStatus.FAILED:
            for col in range(1, len(REPORT_COLUMNS) + 1):
                sheet.cell(row=sheet.max_row, column=col).fill = _FAILED_FILL

    _style(sheet, widths=[8, 34, 26, 34, 90, 70], center_columns=[1])


def _write_details_sheet(sheet: Worksheet, rows: list[ResultRow]) -> None:
    headers = [
        "Sr. No",
        "FileName",
        "Status",
        "Theme",
        "Specific Issue",
        "Reason for Issue",
        "Confidence",
        "Language",
        "Duration (s)",
        "Evidence",
        "Original Transcription",
        "Error",
    ]
    sheet.append(headers)

    for row in rows:
        sheet.append(
            [
                row.sr_no,
                clean_cell(row.file_name),
                row.status.value,
                clean_cell(row.theme),
                clean_cell(row.specific_issue),
                clean_cell(row.reason_for_issue),
                round(row.confidence, 2),
                row.language,
                round(row.duration_seconds, 1),
                clean_cell("\n".join(f'- "{e}"' for e in row.evidence)),
                clean_cell(row.original_transcription),
                clean_cell(row.error or ""),
            ]
        )
        # The source transcript is usually Arabic - render it right to left.
        sheet.cell(row=sheet.max_row, column=11).alignment = _RTL

    _style(
        sheet,
        widths=[8, 34, 14, 24, 30, 45, 12, 10, 12, 45, 80, 40],
        center_columns=[1, 3, 7, 8, 9],
        skip_alignment_columns=[11],
    )


def _write_summary_sheet(
    sheet: Worksheet, job: JobRecord, rows: list[ResultRow]
) -> None:
    successful = [r for r in rows if r.status is CallStatus.COMPLETED]

    sheet.append(["Theme Analytics - Batch Summary"])
    sheet["A1"].font = Font(bold=True, size=14)
    sheet.append([])
    for label, value in (
        ("Job ID", job.id),
        ("Job name", job.name),
        ("Created (UTC)", job.created_at.strftime("%Y-%m-%d %H:%M:%S")),
        ("Status", job.status.value),
        ("Total calls", job.total),
        ("Classified", len(successful)),
        ("Failed", job.failed),
    ):
        sheet.append([label, value])
        sheet.cell(row=sheet.max_row, column=1).font = Font(bold=True)

    sheet.append([])
    header_row = sheet.max_row + 1
    sheet.append(["Theme", "Specific Issue", "Calls", "% of classified"])
    for col in range(1, 5):
        cell = sheet.cell(row=header_row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT

    counts = Counter((r.theme, r.specific_issue) for r in successful)
    total = len(successful) or 1
    for (theme, issue), count in counts.most_common():
        sheet.append([theme, issue, count, f"{count / total * 100:.1f}%"])

    for index, width in enumerate([32, 42, 10, 16], start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def _style(
    sheet: Worksheet,
    widths: list[int],
    center_columns: list[int] | None = None,
    skip_alignment_columns: list[int] | None = None,
) -> None:
    center_columns = center_columns or []
    skip_alignment_columns = skip_alignment_columns or []

    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    for cell in sheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER
    sheet.row_dimensions[1].height = 28

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = _BORDER
            if cell.column in skip_alignment_columns:
                continue
            cell.alignment = _CENTER if cell.column in center_columns else _WRAP

    sheet.freeze_panes = "A2"
    if sheet.max_row > 1:
        sheet.auto_filter.ref = (
            f"A1:{get_column_letter(len(widths))}{sheet.max_row}"
        )
