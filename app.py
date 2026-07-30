
import os
import logging
import pandas as pd
import numpy as np
from datetime import datetime
from dotenv import load_dotenv
import whisper
import wave
import numpy as np
import librosa
import json
import logging
import torch
from silero_vad import load_silero_vad, get_speech_timestamps
from dotenv import load_dotenv
import subprocess
from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import Alignment

load_dotenv()

logger = logging.getLogger(__name__)

whisper_model_path = "/data0/genaiadm_bkp/GenAi-LLM/whisper/large-v3.pt"
whisper_device = torch.device("cuda:0")

torch.cuda.empty_cache()
# output_folder = os.getenv("output_folder")
# base_directory = os.getenv("base_directory")
#log_folder = os.getenv("log_folder")
# pcm_folder = os.getenv("pcm_folder")
#SUMMARISED_DIR = os.getenv("SUMMARISED_DIR")
#DETAILED_DIR = os.getenv("DETAILED_DIR")
#prompts_folder = os.getenv("PROMPT_FOLDER")
INPUT_FOLDER = "./calls"
OUTPUT_FOLDER = "./output"
previous_processed = "./call"

sys_prompt = """You are given a Saudi Awwal Bank's (SAB) customer care conversation transcript generated from a WAV audio file using Whisper.
The transcript is in SRT-like format with precise timestamps for each sentence.

Each line starts with a timestamp in the format:
HH:MM:SSS:MsMs -- HH:MM:SSS:MsMs <speaker_label> <text>

Please translate it to English 

Each output line MUST follow this format:

HH:MM:SSS:MsMs -- HH:MM:SSS:MsMs Agent: <clean English sentence>
or
HH:MM:SSS:MsMs -- HH:MM:SSS:MsMs Customer: <clean English sentence>


7. IMPORTANT CONSTRAINTS

- Do NOT add new content.
- Do NOT summarize.
- Do NOT explain your reasoning.
- Do NOT invent dialogue.
- Only clean, translate, diarize, and mask what is present.
"""

#initial_prompt = "ط¨ظ†ظƒ ط³ط§ط¨ ط§ظ„ظˆط·ظ†ظٹطŒ ط­ط³ط§ط¨طŒ ط¨ط·ط§ظ‚ط©طŒ طھط­ظˆظٹظ„طŒ ط±طµظٹط¯طŒ ط±ظ‚ظ… ط§ظ„ط¢ظٹط¨ط§ظ†طŒ ط®ط¯ظ…ط© ط§ظ„ط¹ظ…ظ„ط§ط،."
#initial_prompt = "SAMA , SAB, Saudi Awwal Bank, Saudi Central Bank, The Saudi Arabian Monetary Agency, ظ…ط¤ط³ط³ط© ط§ظ„ظ†ظ‚ط¯ ط§ظ„ط¹ط±ط¨ظٹ ط§ظ„ط³ط¹ظˆط¯ظٹطŒ ط³ط§ط¨طŒ ط§ظ„ط¨ظ†ظƒ ط§ظ„ط³ط¹ظˆط¯ظٹ ط§ظ„ط£ظˆظ„طŒ ط§ظ„ط¨ظ†ظƒ ط§ظ„ظ…ط±ظƒط²ظٹ ط§ظ„ط³ط¹ظˆط¯ظٹطŒ ظ…ط¤ط³ط³ط© ط§ظ„ظ†ظ‚ط¯ ط§ظ„ط¹ط±ط¨ظٹ ط§ظ„ط³ط¹ظˆط¯ظٹ"
initial_prompt = ""
load_dotenv()

ffmpeg_path = "/opt/genaiadm/call_quality/ffmpeg/ffmpeg"

class WhisperTranscriber:
    def __init__(
        self,
        model_name=whisper_model_path,
        chunk_duration=30,
        silence_pad=0.3,
        target_sample_rate=16000,
        initial_prompt=initial_prompt
    ):
        self.model = whisper.load_model(model_name, device=whisper_device)
        self.vad_model = load_silero_vad()

        self.chunk_duration = chunk_duration
        self.silence_pad = silence_pad
        self.target_sample_rate = target_sample_rate
        self.initial_prompt = initial_prompt

        logger.info("Model Successfully loaded")

    # ---------------- UTILS ----------------
    def format_time(self, seconds):
        hours = int(seconds // 3600)
        minutes = int((seconds % 3600) // 60)
        secs = seconds % 60
        return f"{hours:02d}:{minutes:02d}:{secs:06.2f}"

    # ---------------- AUDIO ----------------
    def load_audio(self, wav_path):
        with wave.open(wav_path, 'rb') as wav:
            sr = wav.getframerate()
            n_channels = wav.getnchannels()
            frames = wav.readframes(wav.getnframes())

        audio_np = np.frombuffer(frames, dtype=np.int16)

        if n_channels > 1:
            audio_np = audio_np.reshape(-1, n_channels).mean(axis=1).astype(np.int16)

        audio_np = audio_np.astype(np.float32) / 32768.0

        if sr != self.target_sample_rate:
            audio_np = librosa.resample(audio_np, orig_sr=sr, target_sr=self.target_sample_rate)

        audio_np = (audio_np * 32768.0).astype(np.int16)
        total_duration_seconds = len(audio_np) / self.target_sample_rate

        return audio_np, total_duration_seconds

    # ---------------- VAD SILENCE LOGIC ----------------
    def get_full_speech_timestamps(self, audio_np):
        audio_float = audio_np.astype(np.float32) / 32768.0

        return get_speech_timestamps(
            audio_float,
            self.vad_model,
            sampling_rate=self.target_sample_rate
        )

    def get_silence_segments_full_audio(self, audio_np, speech_timestamps):
        silence_segments = []
        prev_end = 0

        for seg in speech_timestamps:
            if seg['start'] > prev_end:
                silence_segments.append((prev_end, seg['start']))
            prev_end = seg['end']

        if prev_end < len(audio_np):
            silence_segments.append((prev_end, len(audio_np)))

        return silence_segments

    def merge_silences(self, silence_segments, max_gap_sec=1.0):
        if not silence_segments:
            return []

        merged = []
        cur_start, cur_end = silence_segments[0]

        for start, end in silence_segments[1:]:
            gap = (start - cur_end) / self.target_sample_rate

            if gap <= max_gap_sec:
                cur_end = end
            else:
                merged.append((cur_start, cur_end))
                cur_start, cur_end = start, end

        merged.append((cur_start, cur_end))
        return merged

    def compute_silence_flags(self, audio_np, speech_timestamps):
        silence_segments = self.get_silence_segments_full_audio(audio_np, speech_timestamps)
        silence_segments = self.merge_silences(silence_segments)

        five_sec_flag = True
        thirty_sec_flag = False
        two_min_flag = False

        for start, end in silence_segments:
            duration = (end - start) / self.target_sample_rate
            start_sec = start / self.target_sample_rate

            # Check first 5 sec silence
            if start_sec <= 0:
                if duration < 5:
                    five_sec_flag = False

            # 30 sec silence anywhere
            if duration >= 30:
                thirty_sec_flag = True

            # 2 min silence anywhere
            if duration >= 120:
                two_min_flag = True

        return {
            "5_sec_flag": five_sec_flag,
            "30_sec_flag": thirty_sec_flag,
            "2_min_flag": two_min_flag
        }

    # ---------------- CHUNKING ----------------
    def create_chunks_from_speech(self, audio_np, speech_timestamps):
        """
        Group consecutive VAD speech segments into chunks of roughly
        chunk_duration seconds, but only ever split at a silence gap
        between two speech segments (never mid-word). Each chunk is
        trimmed to its speech span plus a small silence_pad, so leading/
        trailing dead air isn't handed to Whisper.
        """
        if not speech_timestamps:
            return [], []

        pad = int(self.silence_pad * self.target_sample_rate)
        chunk_size_samples = int(self.chunk_duration * self.target_sample_rate)
        audio_len = len(audio_np)

        chunks = []
        time_ranges = []

        def flush(g_start, g_end):
            s = max(0, g_start - pad)
            e = min(audio_len, g_end + pad)
            chunks.append(audio_np[s:e])
            time_ranges.append((s / self.target_sample_rate, e / self.target_sample_rate))

        group_start = speech_timestamps[0]['start']
        group_end = speech_timestamps[0]['end']

        for seg in speech_timestamps[1:]:
            if seg['end'] - group_start > chunk_size_samples:
                flush(group_start, group_end)
                group_start, group_end = seg['start'], seg['end']
            else:
                group_end = seg['end']

        flush(group_start, group_end)

        return chunks, time_ranges

    # ---------------- TRANSCRIPTION ----------------
    def transcribe_chunk(self, chunk, chunk_start_time):
        chunk_float = chunk.astype(np.float32) / 32768.0

        result = self.model.transcribe(
            chunk_float,
            verbose=False,
            language="ar",
            beam_size=5,
            best_of=5,
            temperature=(0.0, 0.2, 0.4, 0.6, 0.8, 1.0),
            condition_on_previous_text=False,
            initial_prompt=self.initial_prompt,
            compression_ratio_threshold=2.4,
            logprob_threshold=-1.0,
            no_speech_threshold=0.6,
        )

        segments = []
        for seg in result.get("segments", []):
            if seg.get("no_speech_prob", 0) > 0.6:
                continue
            if seg.get("avg_logprob", 0) < -1.0:
                continue
            if seg.get("compression_ratio", 0) > 2.4:
                continue

            segments.append({
                "start": self.format_time(chunk_start_time + seg["start"]),
                "end": self.format_time(chunk_start_time + seg["end"]),
                "text": seg["text"].strip()
            })

        main_transcription = " ".join(s["text"] for s in segments) if segments else "[No speech detected]"

        return {
            "main_transcription": main_transcription,
            "segments": segments
        }

    # ---------------- MAIN PIPELINE ----------------
    def transcribe(self, wav_path):
        audio_np, total_duration_seconds = self.load_audio(wav_path)

        speech_timestamps = self.get_full_speech_timestamps(audio_np)
        silence_flags = self.compute_silence_flags(audio_np, speech_timestamps)

        chunks_data, chunk_time_ranges = self.create_chunks_from_speech(audio_np, speech_timestamps)

        file_data = {
            "file": wav_path,
            "total_duration": self.format_time(total_duration_seconds),
            "silence_flags": silence_flags,
            "chunks": []
        }

        if not chunks_data:
            file_data["chunks"].append({
                "start": self.format_time(0),
                "end": self.format_time(total_duration_seconds),
                "main_transcription": "[No speech detected]",
                "segments": []
            })
            return file_data, silence_flags

        for i, chunk in enumerate(chunks_data):
            start_sec, end_sec = chunk_time_ranges[i]

            logger.info(f"Processing chunk {i+1}/{len(chunks_data)}")

            chunk_transcription = self.transcribe_chunk(chunk, start_sec)

            file_data["chunks"].append({
                "start": self.format_time(start_sec),
                "end": self.format_time(end_sec),
                "main_transcription": chunk_transcription["main_transcription"],
                "segments": chunk_transcription["segments"]
            })

        return file_data, silence_flags

    def save_to_json(self, transcription_data, json_path):
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(transcription_data, f, ensure_ascii=False, indent=4)

class JSONToSRTTranslator:
    def __init__(self, prompt_file="prompts/translation_prompt.yaml"):
        self.prompt_file = prompt_file
        self.json_file = None  # Initialize as None; we'll set it later
        self.data = None
        self.whole_transcript = ""
        self.srt_content = []
        self.final_translation = ""
    
    def set_input_file(self, json_file):
        """Set the JSON file and load its data, resetting internal state."""
        self.json_file = json_file
        self.data = self.load_json()
        self.whole_transcript = ""
        self.srt_content = []
        self.final_translation = ""    
        
    def load_json(self):
        """Load JSON data from the provided file."""
        logger.info(f"loading the json file: {self.json_file}")
        with open(self.json_file, "r", encoding="utf-8") as file:
            return json.load(file)
    
    def format_time(self, timestamp):
        """Format time into SRT format (hh:mm:ss,ms)."""
        parts = timestamp.split(":")
        seconds, milliseconds = parts[2].split(".")
        return f"{parts[0]}:{parts[1]}:{seconds}.{milliseconds[:3]}"
    
    def convert_to_srt(self):
        """Convert JSON transcript into SRT format."""
        logger.info("Convert JSON transcript into SRT format.")
        for chunk in self.data.get("chunks", []):
            self.whole_transcript += chunk.get("main_transcription", "") + "\n"
            for segment in chunk.get("segments", []):
                start_time = self.format_time(segment["start"])
                end_time = self.format_time(segment["end"])
                text = segment.get("text", "")
                self.srt_content.append(f"{start_time} -- {end_time} : {text}\n")
    
    def check_no_speech(self):
        chunks = self.data.get("chunks", [])
        speech_txt = chunks[0].get("main_transcription", "")
        if speech_txt == "[No speech detected]":
            return True
        else:
            return False
                
    def load_prompt_from_yaml(self, **kwargs):
        """Load and format the prompt from a YAML file."""
        try:
            with open(self.prompt_file, 'r') as file:
                prompt_data = yaml.safe_load(file)
        except FileNotFoundError:
            logger.error(f"Prompt file not found: {self.prompt_file}")
            raise
        
        system_message = prompt_data['system_message']
        user_message = prompt_data['user_message'].format(**kwargs)
        return [
            {"role": "system", "content": system_message},
            {"role": "user", "content": user_message}
        ]
    def generate_prompt(self):
        """Generate the prompt for translation using the YAML file."""
        return self.load_prompt_from_yaml(
            whole_transcript=self.whole_transcript,
            srt_context="".join(self.srt_content)
        )
    
    def translate(self):
        """Send the prompt to the OpenAI API for translation and stream the output."""
        srt_context="".join(self.srt_content)
        openai_api_key = "EMPTY"
        openai_api_base = "http://localhost:8003/v1"
        client = OpenAI(api_key=openai_api_key, base_url=openai_api_base)
        # Get model ID from available models
        models = client.models.list()
        model = models.data[0].id
        print("\nMODEL:::",model)
        logger.info("Started translating")
        # stream = client.chat.completions.create(model=model, 
        #                                         messages=self.generate_prompt(), 
        #                                         # extra_body={"guided_json": json_schema},
        #                                         stream=True)
        stream = client.chat.completions.create(model=model,
                                                temperature=0.1,
                                                messages=[{"role": "system", "content": [{"type": "text", "text": sys_prompt}]},
                                                            {"role": "user", "content": [{"type": "text", "text": f"PROCESS THE FOLLOWING TRANSCRIPT:::\n{srt_context}"}]}], 
                                                # extra_body={"guided_json": json_schema},
                                                stream=True,
                                                max_tokens=7000)
        for chunk in stream:
            # Depending on how the delta is returned, extract content accordingly
            delta = chunk.choices[0].delta
            if hasattr(delta, "content") and delta.content is not None:
                self.final_translation += delta.content
    

    def save_translation(self, output_file="final_translation.txt"):
        """Save the translated output to a file."""
        with open(output_file, "w", encoding="utf-8") as outfile:
            outfile.write(self.final_translation)
        # logger.info(f"Translation saved to {output_file}")

transcriber = WhisperTranscriber()

translator = JSONToSRTTranslator()


def convert_single_audio(input_path, output_folder):
    """
    Convert audio to PCM WAV and save in the output folder.
    """

    os.makedirs(output_folder, exist_ok=True)
    base_name = os.path.splitext(os.path.basename(input_path))[0]

    output_path = os.path.join(
        output_folder,
        f"{base_name}_pcm.wav"
    )

    logger.info(f"Converting: {input_path}")

    cmd = [
        ffmpeg_path,
        "-y",
        "-i",
        input_path,
        "-acodec",
        "pcm_s16le",
        "-ar",
        "16000",
        "-ac",
        "1",
        output_path
    ]

    subprocess.run(cmd, check=True)

    logger.info(f"PCM saved: {output_path}")

    return output_path


def build_transcript_paragraph(transcription_data):
    """Join every chunk's transcription into a single flowing paragraph."""
    parts = [
        chunk["main_transcription"]
        for chunk in transcription_data.get("chunks", [])
        if chunk.get("main_transcription") and chunk["main_transcription"] != "[No speech detected]"
    ]
    return " ".join(parts).strip()


def save_results_to_excel(results, excel_path):
    """Write filename/transcription/translation rows, with the Arabic
    transcription column rendered right-to-left."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Transcripts"

    ws.append(["filename", "transcription", "translation"])

    ltr_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    rtl_alignment = Alignment(horizontal="right", vertical="top", wrap_text=True, readingOrder=2)

    for row in results:
        ws.append([row["filename"], row["transcription"], row["translation"]])
        r = ws.max_row
        ws.cell(row=r, column=1).alignment = ltr_alignment
        ws.cell(row=r, column=2).alignment = rtl_alignment
        ws.cell(row=r, column=3).alignment = ltr_alignment

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 80
    ws.column_dimensions["C"].width = 80

    wb.save(excel_path)


def run_pipeline_single(audio_file, transcriber, translator, output_folder):
    try:
        logger.info(f"Processing {audio_file}")

        os.makedirs(output_folder, exist_ok=True)

        base_name = os.path.splitext(
            os.path.basename(audio_file)
        )[0]

        transcription_json = os.path.join(
            output_folder,
            f"{base_name}_transcription.json"
        )

        translation_txt = os.path.join(
            output_folder,
            f"{base_name}_translation.txt"
        )

        # Convert audio
        pcm_audio_file = convert_single_audio(
            audio_file,
            output_folder
        )

        # Transcribe
        transcription_data, silence_flags = (
            transcriber.transcribe(
                pcm_audio_file
            )
        )

        transcriber.save_to_json(
            transcription_data,
            transcription_json
        )

        logger.info(
            f"Saved transcription: {transcription_json}"
        )

        # Translate
        translator.set_input_file(
            transcription_json
        )

        translator.convert_to_srt()
        translator.translate()
        translator.save_translation(
            translation_txt
        )

        logger.info(
            f"Saved translation: {translation_txt}"
        )

        return {
            "status": "success",
            "transcription": transcription_json,
            "translation": translation_txt,
            "silence_flags": silence_flags,
            "transcript_text": build_transcript_paragraph(transcription_data),
            "translation_text": translator.final_translation
        }

    except Exception as e:
        logger.exception("Pipeline failed")
        return {
            "status": "error",
            "message": str(e)
        }


if __name__ == "__main__":

    audio_extensions = (
        ".wav",
    )

    audio_files = [
        os.path.join(INPUT_FOLDER, f)
        for f in os.listdir(INPUT_FOLDER)
        if f.lower().endswith(audio_extensions) and not f.lower().endswith("_pcm.wav")
        and f not in previous_processed
    ]

    if len(audio_files) == 0:
        raise Exception(
            f"No audio file found in {INPUT_FOLDER}"
        )

    os.makedirs(OUTPUT_FOLDER, exist_ok=True)

    results = []
    excel_rows = []

    for audio_file in audio_files:
        result = run_pipeline_single(
            audio_file,
            transcriber,
            translator,
            OUTPUT_FOLDER
        )
        results.append({"file": audio_file, **result})

        excel_rows.append({
            "filename": os.path.basename(audio_file),
            "transcription": result.get("transcript_text", ""),
            "translation": result.get("translation_text", "") if result["status"] == "success" else f"ERROR: {result.get('message', '')}"
        })

    excel_path = os.path.join(OUTPUT_FOLDER, "transcripts_summary.xlsx")
    save_results_to_excel(excel_rows, excel_path)
    logger.info(f"Saved Excel summary: {excel_path}")

    print(json.dumps(results, indent=4))